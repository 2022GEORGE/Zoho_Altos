from django.shortcuts import render,redirect
from Register_Login.models import *
from Register_Login.views import logout
from django.contrib import messages
from django.conf import settings
from datetime import date
from datetime import datetime, timedelta
from Company_Staff.models import *
import os
import pandas as pd
from django.http import JsonResponse,HttpResponse,HttpResponseRedirect
from openpyxl import load_workbook
from django.core.mail import send_mail, EmailMessage
from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa
# Create your views here.



# -------------------------------Company section--------------------------------

# company dashboard
def company_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        # Calculate the date 20 days before the end date for payment term renew
        reminder_date = dash_details.End_date - timedelta(days=20)
        current_date = date.today()
        alert_message = current_date >= reminder_date
        
        message='.'
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'alert_message':alert_message
        }
        return render(request, 'company/company_dash.html', context)
    else:
        return redirect('/')


# company staff request for login approval
def company_staff_request(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        staff_request=StaffDetails.objects.filter(company=dash_details.id, company_approval=0).order_by('-id')
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'requests':staff_request,
        }
        return render(request, 'company/staff_request.html', context)
    else:
        return redirect('/')

# company staff accept or reject
def staff_request_accept(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    staff.company_approval=1
    staff.save()
    return redirect('company_staff_request')

def staff_request_reject(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    login_details=LoginDetails.objects.get(id=staff.company.id)
    login_details.delete()
    staff.delete()
    return redirect('company_staff_request')


# All company staff view, cancel staff approval
def company_all_staff(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        all_staffs=StaffDetails.objects.filter(company=dash_details.id, company_approval=1).order_by('-id')
       
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'staffs':all_staffs,
        }
        return render(request, 'company/all_staff_view.html', context)
    else:
        return redirect('/')

def staff_approval_cancel(request, pk):
    """
    Sets the company approval status to 2 for the specified staff member, effectively canceling staff approval.

    This function is designed to be used for canceling staff approval, and the company approval value is set to 2.
    This can be useful for identifying resigned staff under the company in the future.

    """
    staff = StaffDetails.objects.get(id=pk)
    staff.company_approval = 2
    staff.save()
    return redirect('company_all_staff')


# company profile, profile edit
def company_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        terms=PaymentTerms.objects.all()

        # Calculate the date 20 days before the end date
        reminder_date = dash_details.End_date - timedelta(days=20)
        current_date = date.today()
        renew_button = current_date >= reminder_date
        
        
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'renew_button': renew_button,
            'terms':terms,
        }
        return render(request, 'company/company_profile.html', context)
    else:
        return redirect('/')

def company_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_profile_editpage.html', context)
    else:
        return redirect('/')

def company_profile_basicdetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            messages.success(request,'Updated')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
    
def company_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                log_details.password=password
                log_details.save()

            messages.success(request,'Password Changed')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
       
def company_profile_companydetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            gstno = request.POST.get('gstno')
            profile_pic = request.FILES.get('image')

            # Update the CompanyDetails object with form data
            dash_details.company_name = request.POST.get('cname')
            dash_details.contact = request.POST.get('phone')
            dash_details.address = request.POST.get('address')
            dash_details.city = request.POST.get('city')
            dash_details.state = request.POST.get('state')
            dash_details.country = request.POST.get('country')
            dash_details.pincode = request.POST.get('pincode')
            dash_details.pan_number = request.POST.get('pannumber')

            if gstno:
                dash_details.gst_no = gstno

            if profile_pic:
                dash_details.profile_pic = profile_pic

            dash_details.save()

            messages.success(request, 'Updated')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/')    

# company modules editpage
def company_module_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_module_editpage.html', context)
    else:
        return redirect('/')

def company_module_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        if request.method == 'POST':
            # Retrieve values
            items = request.POST.get('items', 0)
            price_list = request.POST.get('price_list', 0)
            stock_adjustment = request.POST.get('stock_adjustment', 0)
            godown = request.POST.get('godown', 0)

            cash_in_hand = request.POST.get('cash_in_hand', 0)
            offline_banking = request.POST.get('offline_banking', 0)
            upi = request.POST.get('upi', 0)
            bank_holders = request.POST.get('bank_holders', 0)
            cheque = request.POST.get('cheque', 0)
            loan_account = request.POST.get('loan_account', 0)

            customers = request.POST.get('customers', 0)
            invoice = request.POST.get('invoice', 0)
            estimate = request.POST.get('estimate', 0)
            sales_order = request.POST.get('sales_order', 0)
            recurring_invoice = request.POST.get('recurring_invoice', 0)
            retainer_invoice = request.POST.get('retainer_invoice', 0)
            credit_note = request.POST.get('credit_note', 0)
            payment_received = request.POST.get('payment_received', 0)
            delivery_challan = request.POST.get('delivery_challan', 0)

            vendors = request.POST.get('vendors', 0)
            bills = request.POST.get('bills', 0)
            recurring_bills = request.POST.get('recurring_bills', 0)
            vendor_credit = request.POST.get('vendor_credit', 0)
            purchase_order = request.POST.get('purchase_order', 0)
            expenses = request.POST.get('expenses', 0)
            recurring_expenses = request.POST.get('recurring_expenses', 0)
            payment_made = request.POST.get('payment_made', 0)

            projects = request.POST.get('projects', 0)

            chart_of_accounts = request.POST.get('chart_of_accounts', 0)
            manual_journal = request.POST.get('manual_journal', 0)

            eway_bill = request.POST.get('ewaybill', 0)

            employees = request.POST.get('employees', 0)
            employees_loan = request.POST.get('employees_loan', 0)
            holiday = request.POST.get('holiday', 0)
            attendance = request.POST.get('attendance', 0)
            salary_details = request.POST.get('salary_details', 0)

            reports = request.POST.get('reports', 0)

            update_action=1
            status='Pending'

            # Create a new ZohoModules instance and save it to the database
            data = ZohoModules(
                company=dash_details,
                items=items, price_list=price_list, stock_adjustment=stock_adjustment, godown=godown,
                cash_in_hand=cash_in_hand, offline_banking=offline_banking, upi=upi, bank_holders=bank_holders,
                cheque=cheque, loan_account=loan_account,
                customers=customers, invoice=invoice, estimate=estimate, sales_order=sales_order,
                recurring_invoice=recurring_invoice, retainer_invoice=retainer_invoice, credit_note=credit_note,
                payment_received=payment_received, delivery_challan=delivery_challan,
                vendors=vendors, bills=bills, recurring_bills=recurring_bills, vendor_credit=vendor_credit,
                purchase_order=purchase_order, expenses=expenses, recurring_expenses=recurring_expenses,
                payment_made=payment_made,
                projects=projects,
                chart_of_accounts=chart_of_accounts, manual_journal=manual_journal,
                eway_bill=eway_bill,
                employees=employees, employees_loan=employees_loan, holiday=holiday,
                attendance=attendance, salary_details=salary_details,
                reports=reports,update_action=update_action,status=status    
            )
            data.save()
            messages.info(request,'Request sent successfully, wait for approval...')
            return redirect('company_module_editpage')
        else:
            return redirect('company_module_editpage')  
    else:
        return redirect('/')


def company_renew_terms(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        if request.method == 'POST':
            select=request.POST['select']
            terms=PaymentTerms.objects.get(id=select)
            update_action=1
            status='Pending'
            newterms=PaymentTermsUpdates(
               company=dash_details,
               payment_term=terms,
               update_action=update_action,
               status=status 
            )
            newterms.save()
            messages.info(request,'Successfully requested for extend payment terms, please wait for approval.')
            return redirect('company_profile')
    else:
        return redirect('/')









# -------------------------------Staff section--------------------------------

# staff dashboard
def staff_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_dash.html',context)
    else:
        return redirect('/')


# staff profile
def staff_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_profile.html',context)
    else:
        return redirect('/')


def staff_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'staff/staff_profile_editpage.html', context)
    else:
        return redirect('/')

def staff_profile_details_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            dash_details.contact = request.POST.get('phone')
            old=dash_details.image
            new=request.FILES.get('profile_pic')
            print(new,old)
            if old!=None and new==None:
                dash_details.image=old
            else:
                print(new)
                dash_details.image=new
            dash_details.save()
            messages.success(request,'Updated')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')

def staff_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                log_details.password=password
                log_details.save()

            messages.success(request,'Password Changed')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')






# -------------------------------Zoho Modules section--------------------------------
#------------------------------------payroll employee--------------------------------
#------------------------------------------------GEORGE MATHEW---------------------------------------
def payroll_employee_create(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    blood=Bloodgroup.objects.all()
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'log_id':log_details,
            'blood':blood
            
    }
    return render(request,'zohomodules/payroll-employee/payroll_create_employee.html',content)
def employee_list(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        pay=payroll_employee.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        content = {
                'details': dash_details,
                'pay':pay,
                'allmodules': allmodules,
                'log_id':log_details
        }
        return render(request,'zohomodules/payroll-employee/payroll_list.html',content)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        pay=payroll_employee.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        content = {
                'details': dash_details,
                'pay':pay,
                'allmodules': allmodules,
                'log_id':log_details
        }
        return render(request,'zohomodules/payroll-employee/payroll_list.html',content)
def employee_overview(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type =='Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        pay=payroll_employee.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        p=payroll_employee.objects.get(id=pk)
        comment_data=comment.objects.filter(login_details=log_details,employee=pk)
        history=employee_history.objects.filter(login_details=log_details,employee=pk)
    if log_details.user_type =='Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        pay=payroll_employee.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        p=payroll_employee.objects.get(id=pk)
        comment_data=comment.objects.filter(login_details=log_details,employee=pk)
        history=employee_history.objects.all()
    content = {
                'details': dash_details,
                'pay':pay,
                'p':p,
                'allmodules': allmodules,
                'comment':comment_data,
                'history':history,
                'log_id':log_details,
        }
    return render(request,'zohomodules/payroll-employee/overview_page.html',content)
def create_employee(request):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':    
            company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            title=request.POST['title']
            fname=request.POST['fname']
            lname=request.POST['lname']
            alias=request.POST['alias']
            joindate=request.POST['joindate']
            salarydate=request.POST['salary']
            saltype=request.POST['saltype']
            if (saltype == 'Fixed'):
                salary=request.POST['fsalary']
            else:
                salary=request.POST['vsalary']
            image=request.FILES.get('file')
            amountperhr=request.POST['amnthr']
            workhr=request.POST['hours'] 
            empnum=request.POST['empnum']
            if payroll_employee.objects.filter(emp_number=empnum,company=company_details):
                messages.info(request,'employee number all ready exists')
                return redirect('payroll_employee_create')
            designation = request.POST['designation']
            location=request.POST['location']
            gender=request.POST['gender']
            dob=request.POST['dob']
            blood=request.POST['blood']
            fmname=request.POST['fm_name']
            sname=request.POST['s_name']        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            phone=request.POST['phone']
            ephone=request.POST['ephone']
            result_set1 = payroll_employee.objects.filter(company=company_details,Phone=phone)
            result_set2 = payroll_employee.objects.filter(company=company_details,emergency_phone=ephone)
            if result_set1:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_create')
            if result_set2:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_create')
            email=request.POST['email']
            result_set = payroll_employee.objects.filter(company=company_details,email=email)
            if result_set:
                messages.error(request,'email already exists')
                return redirect('payroll_employee_create')
            isdts=request.POST['tds']
            attach=request.FILES.get('attach')
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']
            an=request.POST['an']
            if payroll_employee.objects.filter(Aadhar=an,company=company_details):
                    messages.error(request,'Aadhra number already exists')
                    return redirect('payroll_employee_create')   
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
                payroll= payroll_employee(title=title,first_name=fname,last_name=lname,alias=alias,image=image,joindate=joindate,salary_type=saltype,salary=salary,age=age,
                            emp_number=empnum,designation=designation,location=location, gender=gender,dob=dob,blood=blood,parent=fmname,spouse_name=sname,workhr=workhr,
                            amountperhr = amountperhr, address=address,permanent_address=paddress ,Phone=phone,emergency_phone=ephone, email=email,Income_tax_no=itn,Aadhar=an,
                            UAN=uan,PFN=pfn,PRAN=pran,uploaded_file=attach,isTDS=istdsval,TDS_percentage=tds,salaryrange = salarydate,acc_no=accno,IFSC=ifsc,bank_name=bname,branch=branch,transaction_type=ttype,company=dash_details,login_details=log_details)
                payroll.save()
                history=employee_history(company=dash_details,login_details=log_details, employee=payroll,Action='CREATED')
                history.save()
                messages.info(request,'employee created')
                return redirect('employee_list')
        if log_details.user_type == 'Staff':
            company_details = StaffDetails.objects.get(login_details=log_details)
            title=request.POST['title']
            fname=request.POST['fname']
            lname=request.POST['lname']
            alias=request.POST['alias']
            joindate=request.POST['joindate']
            salarydate=request.POST['salary']
            saltype=request.POST['saltype']
            if (saltype == 'Fixed'):
                salary=request.POST['fsalary']
            else:
                salary=request.POST['vsalary']
            image=request.FILES.get('file')
            amountperhr=request.POST['amnthr']
            workhr=request.POST['hours'] 
            empnum=request.POST['empnum']
            if payroll_employee.objects.filter(emp_number=empnum,company=company_details.company):
                messages.info(request,'employee number all ready exists')
                return redirect('payroll_employee_create')
            designation = request.POST['designation']
            location=request.POST['location']
            gender=request.POST['gender']
            dob=request.POST['dob']
            blood=request.POST['blood']
            fmname=request.POST['fm_name']
            sname=request.POST['s_name']        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            phone=request.POST['phone']
            ephone=request.POST['ephone']
            result_set1 = payroll_employee.objects.filter(company=company_details.company,Phone=phone)
            result_set2 = payroll_employee.objects.filter(company=company_details.company,emergency_phone=ephone)
            if result_set1:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_create')
            if result_set2:
                messages.error(request,'emerency phone no already exists')
                return redirect('payroll_employee_create')
            email=request.POST['email']
            result_set = payroll_employee.objects.filter(company=company_details.company,email=email)
            if result_set:
                messages.error(request,'email already exists')
                return redirect('payroll_employee_create')
            isdts=request.POST['tds']
            attach=request.FILES.get('attach')
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']
            an=request.POST['an']
            if payroll_employee.objects.filter(Aadhar=an,company=company_details.company):
                    messages.error(request,'Aadhra number already exists')
                    return redirect('payroll_employee_create')   
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            dash_details = StaffDetails.objects.get(login_details=log_details)
            payroll= payroll_employee(title=title,first_name=fname,last_name=lname,alias=alias,image=image,joindate=joindate,salary_type=saltype,salary=salary,age=age,
                         emp_number=empnum,designation=designation,location=location, gender=gender,dob=dob,blood=blood,parent=fmname,spouse_name=sname,workhr=workhr,
                         amountperhr = amountperhr, address=address,permanent_address=paddress ,Phone=phone,emergency_phone=ephone, email=email,Income_tax_no=itn,Aadhar=an,
                         UAN=uan,PFN=pfn,PRAN=pran,uploaded_file=attach,isTDS=istdsval,TDS_percentage=tds,salaryrange = salarydate,acc_no=accno,IFSC=ifsc,bank_name=bname,branch=branch,transaction_type=ttype,company=dash_details.company,login_details=log_details)
            payroll.save()
            history=employee_history(company=dash_details.company,login_details=log_details, employee=payroll,Action='CREATED')
            history.save()
            messages.info(request,'employee created')
            return redirect('employee_list')
    return redirect('payroll_employee_create')
def payroll_employee_edit(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    blood=Bloodgroup.objects.all()
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        p=payroll_employee.objects.get(id=pk)
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        p=payroll_employee.objects.get(id=pk)
        
    print(p)
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'p':p,
            'log_id':log_details,
            'blood':blood
    }
    return render(request,'zohomodules/payroll-employee/edit_employee.html',content)
def do_payroll_edit(request,pk):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type =='Company':
            company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)    
            title=request.POST['title']
            fname=request.POST['fname']
            lname=request.POST['lname']
            alias=request.POST['alias']
            joindate=request.POST['joindate']
            salarydate=request.POST['salary']
            saltype=request.POST['saltype']
            if (saltype == 'Fixed' or saltype =='Temporary'):
                salary=request.POST['fsalary']
            else:
                salary=request.POST['vsalary']
            image=request.FILES.get('file')
            amountperhr=request.POST['amnthr']
            workhr=request.POST['hours']
            empnum=request.POST['empnum']
            result_set2 = payroll_employee.objects.filter(company=company_details,emp_number=empnum).exclude(id=pk)
            if result_set2:
                messages.error(request,'employee number  already exists')
                return redirect('payroll_employee_edit',pk)
            designation = request.POST['designation']
            location=request.POST['location']
            gender=request.POST['gender']
            dob=request.POST['dob']
            blood=request.POST['blood']
            fmname=request.POST['fm_name']
            sname=request.POST['s_name']        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            phone=request.POST['phone']
            ephone=request.POST['ephone']
            result_set1 = payroll_employee.objects.filter(company=company_details,Phone=phone).exclude(id=pk)
            result_set3 = payroll_employee.objects.filter(company=company_details,emergency_phone=phone).exclude(id=pk)
            if result_set1:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_edit',pk)
            if result_set3:
                messages.error(request,'emergency phone no already exists')
                return redirect('payroll_employee_edit',pk)
            email=request.POST['email']
            result_set = payroll_employee.objects.filter(company=company_details,email=email).exclude(id=pk)
            if result_set:
                messages.error(request,'email already exists')
                return redirect('payroll_employee_edit',pk)
            isdts=request.POST['tds']
            attach=request.FILES.get('attach')
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']
            an=request.POST['an'] 
            if payroll_employee.objects.filter(Aadhar=an,company=company_details).exclude(id=pk):
                messages.error(request,'Aadhra number already exists')
                return redirect('payroll_employee_edit',pk)
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
                payroll= payroll_employee.objects.get(id=pk)
                payroll.title=title
                payroll.first_name=fname
                payroll.last_name=lname
                payroll.alias=alias
                if len(request.FILES) != 0:
                    if image :
                        if payroll.image:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.image.path):
                                    os.remove(payroll.image.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.image = image
                        else:
                            # Assign the new file to payroll.image
                            payroll.image = image
                payroll.joindate=joindate
                payroll.salary_type=saltype
                payroll.salary=salary
                age=age
                payroll.emp_number=empnum
                payroll.designation=designation
                payroll.location=location
                payroll.gender=gender
                payroll.dob=dob
                payroll.blood=blood
                payroll.parent=fmname
                payroll.spouse_name=sname
                payroll.workhr=workhr
                payroll.amountperhr = amountperhr
                payroll.address=address
                payroll.permanent_address=paddress
                payroll.Phone=phone
                payroll.emergency_phone=ephone
                payroll.email=email
                payroll.Income_tax_no=itn
                payroll.Aadhar=an
                payroll.UAN=uan
                payroll.PFN=pfn
                payroll.PRAN=pran
                if len(request.FILES) !=0:
                    if attach :
                        if payroll.uploaded_file:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.uploaded_file.path):
                                    os.remove(payroll.uploaded_file.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                        else:
                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                payroll.isTDS=istdsval
                payroll.TDS_percentage=tds
                payroll.salaryrange = salarydate
                payroll.acc_no=accno
                payroll.IFSC=ifsc
                payroll.bank_name=bname
                payroll.branch=branch
                payroll.transaction_type=ttype
                payroll.company=dash_details
                payroll.login_details=log_details
                payroll.save()
                history=employee_history(company=dash_details,login_details=log_details, employee=payroll,Action='EDITED')
                history.save()
                messages.info(request,'Updated')
                return redirect('employee_overview',pk)
        if log_details.user_type == 'Staff':
            if log_details.user_type =='Staff':
                company_details = StaffDetails.objects.get(login_details=log_details)    
                title=request.POST['title']
                fname=request.POST['fname']
                lname=request.POST['lname']
                alias=request.POST['alias']
                joindate=request.POST['joindate']
                salarydate=request.POST['salary']
                saltype=request.POST['saltype']
                if (saltype == 'Fixed' or saltype =='Temporary'):
                    salary=request.POST['fsalary']
                else:
                    salary=request.POST['vsalary']
                image=request.FILES.get('file')
                amountperhr=request.POST['amnthr']
                workhr=request.POST['hours']
                empnum=request.POST['empnum']
                result_set2 = payroll_employee.objects.filter(company=company_details.company,emp_number=empnum).exclude(id=pk)
                if result_set2:
                    messages.error(request,'employee number  already exists')
                    return redirect('payroll_employee_edit',pk)
                designation = request.POST['designation']
                location=request.POST['location']
                gender=request.POST['gender']
                dob=request.POST['dob']
                blood=request.POST['blood']
                fmname=request.POST['fm_name']
                sname=request.POST['s_name']        
                add1=request.POST['address']
                add2=request.POST['address2']
                address=add1+" "+add2
                padd1=request.POST['paddress'] 
                padd2=request.POST['paddress2'] 
                paddress= padd1+padd2
                phone=request.POST['phone']
                ephone=request.POST['ephone']
                result_set1 = payroll_employee.objects.filter(company=company_details.company,Phone=phone).exclude(id=pk)
                result_set3 = payroll_employee.objects.filter(company=company_details.company,emergency_phone=ephone).exclude(id=pk)
                if result_set1:
                    messages.error(request,'phone no already exists')
                    return redirect('payroll_employee_edit',pk)
                if result_set3:
                    messages.error(request,'emergency phone no already exists')
                    return redirect('payroll_employee_edit',pk)
                email=request.POST['email']
                result_set = payroll_employee.objects.filter(company=company_details.company,email=email).exclude(id=pk)
                if result_set:
                    messages.error(request,'email already exists')
                    return redirect('payroll_employee_edit',pk)
                isdts=request.POST['tds']
                attach=request.FILES.get('attach')
                if isdts == '1':
                    istdsval=request.POST['pora']
                    if istdsval == 'Percentage':
                        tds=request.POST['pcnt']
                    elif istdsval == 'Amount':
                        tds=request.POST['amnt']
                else:
                    istdsval='No'
                    tds = 0
                itn=request.POST['itn']
                an=request.POST['an'] 
                if payroll_employee.objects.filter(Aadhar=an,company=company_details.company).exclude(id=pk):
                    messages.error(request,'Aadhra number already exists')
                    return redirect('payroll_employee_edit',pk)
                uan=request.POST['uan'] 
                pfn=request.POST['pfn']
                pran=request.POST['pran']
                age=request.POST['age']
                bank=request.POST['bank']
                accno=request.POST['acc_no']       
                ifsc=request.POST['ifsc']       
                bname=request.POST['b_name']       
                branch=request.POST['branch']
                ttype=request.POST['ttype']
                dash_details = StaffDetails.objects.get(login_details=log_details)
                payroll= payroll_employee.objects.get(id=pk)
                payroll.title=title
                payroll.first_name=fname
                payroll.last_name=lname
                payroll.alias=alias
                if len(request.FILES) != 0:
                    if image :
                        if payroll.image:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.image.path):
                                    os.remove(payroll.image.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.image = image
                        else:
                            # Assign the new file to payroll.image
                            payroll.image = image
                payroll.joindate=joindate
                payroll.salary_type=saltype
                payroll.salary=salary
                age=age
                payroll.emp_number=empnum
                payroll.designation=designation
                payroll.location=location
                payroll.gender=gender
                payroll.dob=dob
                payroll.blood=blood
                payroll.parent=fmname
                payroll.spouse_name=sname
                payroll.workhr=workhr
                payroll.amountperhr = amountperhr
                payroll.address=address
                payroll.permanent_address=paddress
                payroll.Phone=phone
                payroll.emergency_phone=ephone
                payroll.email=email
                payroll.Income_tax_no=itn
                payroll.Aadhar=an
                payroll.UAN=uan
                payroll.PFN=pfn
                payroll.PRAN=pran
                if len(request.FILES) !=0:
                    if attach :
                        if payroll.uploaded_file:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.uploaded_file.path):
                                    os.remove(payroll.uploaded_file.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                        else:
                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                payroll.isTDS=istdsval
                payroll.TDS_percentage=tds
                payroll.salaryrange = salarydate
                payroll.acc_no=accno
                payroll.IFSC=ifsc
                payroll.bank_name=bname
                payroll.branch=branch
                payroll.transaction_type=ttype
                payroll.company=dash_details.company
                payroll.login_details=log_details
                payroll.save()
                history=employee_history(company=dash_details.company,login_details=log_details, employee=payroll,Action='EDITED')
                history.save()
                messages.info(request,'Updated')
                return redirect('employee_overview',pk)
    return redirect('employee_overview',pk)
def add_comment(request,pk):
    if request.method =='POST':
        comment_data=request.POST['comments']
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        payroll= payroll_employee.objects.get(id=pk) 
        data=comment(comment=comment_data,login_details=log_details,employee=payroll)
        data.save()
        return redirect('employee_overview',pk)
    return redirect('employee_overview',pk)
def delete_commet(request,pk,pi):
    data=comment.objects.get(id=pk)
    data.delete()
    return redirect('employee_overview',pi)
def delete_employee(request,pk):
    data=payroll_employee.objects.get(id=pk)
    data.delete()
    return redirect('employee_list')
def employee_status(request,pk):
    data=payroll_employee.objects.get(id=pk)
    if data.status == 'Active':
        data.status ='Inactive'
    elif data.status == 'Inactive':
        data.status ='Active'
    data.save()
    return redirect('employee_overview',pk)
def add_blood(request):
    if request.method == 'POST':
        blood = request.POST.get('blood')
        print(blood)

        # Validate input
        if not blood:
            return JsonResponse({'message': 'Invalid or missing blood group'})

        # Use get_or_create for simplicity
        if Bloodgroup.objects.filter(Blood_group=blood):
            return JsonResponse({'message': 'Blood group already exists'})
        Bloodgroup.objects.create(Blood_group=blood)
        data=Bloodgroup.objects.all()
        return JsonResponse({'message': 'Blood group added','blood' : blood})
def import_payroll_excel(request):
    print(1)
    print('hello')
    if request.method == 'POST' :
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(2, eb.max_row + 1):
                    billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    payroll=payroll_employee(title=billsheet[0],first_name=billsheet[1],last_name=billsheet[2],alias=billsheet[3],joindate=datetime.date.fromisoformat(billsheet[4]),salary_type=billsheet[6],salary=billsheet[9],
                                emp_number=billsheet[10],designation=billsheet[11],location=billsheet[12], gender=billsheet[13],dob=datetime.date.fromisoformat(billsheet[14]),blood=billsheet[15],parent=billsheet[16],spouse_name=billsheet[17],workhr=billsheet[8],
                                amountperhr = billsheet[7], address=billsheet[19],permanent_address=billsheet[18],Phone=billsheet[20],emergency_phone=billsheet[21], email=billsheet[22],Income_tax_no=billsheet[32],Aadhar=billsheet[33],
                                UAN=billsheet[34],PFN=billsheet[35],PRAN=billsheet[36],isTDS=billsheet[29],TDS_percentage=billsheet[30],salaryrange = billsheet[5],acc_no=billsheet[24],IFSC=billsheet[25],bank_name=billsheet[26],branch=billsheet[27],transaction_type=billsheet[28],company=dash_details.company,login_details=log_details)
                    payroll.save()
                    history=employee_history(company=dash_details.company,login_details=log_details, employee=payroll,Action='IMPORTED')
                    history.save()
                    messages.warning(request,'file imported')
                    return redirect('employee_list')
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(2, eb.max_row + 1):
                    billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    payroll=payroll_employee(title=billsheet[0],first_name=billsheet[1],last_name=billsheet[2],alias=billsheet[3],joindate=billsheet[4],salary_type=billsheet[6],salary=billsheet[9],
                                emp_number=billsheet[10],designation=billsheet[11],location=billsheet[12], gender=billsheet[13],dob=billsheet[14],blood=billsheet[15],parent=billsheet[16],spouse_name=billsheet[17],workhr=billsheet[8],
                                amountperhr = billsheet[7], address=billsheet[19],permanent_address=billsheet[18],Phone=billsheet[20],emergency_phone=billsheet[21], email=billsheet[22],Income_tax_no=billsheet[32],Aadhar=billsheet[33],
                                UAN=billsheet[34],PFN=billsheet[35],PRAN=billsheet[36],isTDS=billsheet[29],TDS_percentage=billsheet[30],salaryrange = billsheet[5],acc_no=billsheet[24],IFSC=billsheet[25],bank_name=billsheet[26],branch=billsheet[27],transaction_type=billsheet[28],company=dash_details,login_details=log_details)
                    payroll.save()
                    history=employee_history(company=dash_details,login_details=log_details, employee=payroll,Action='IMPORTED')
                    history.save()
                    messages.warning(request,'file imported')
                    return redirect('employee_list')
    messages.error(request,'File upload Failed!11')
    return redirect('employee_list')
def add_file(request,pk):
    if request.method == 'POST':
        data=request.FILES.get('file')
        payroll=payroll_employee.objects.get(id=pk)
        if payroll.uploaded_file:
            try:
                                # Check if the file exists before removing it
                if os.path.exists(payroll.uploaded_file.path):
                    os.remove(payroll.uploaded_file.path)
            except Exception as e:
                messages.error(request,'file upload error')
                return redirect('employee_overview',pk)

                            # Assign the new file to payroll.image
            payroll.uploaded_file = data
            payroll.save()
            messages.info(request,'fil uploaded')
            return redirect('employee_overview',pk)
        else:
            payroll.uploaded_file = data
            payroll.save()
        messages.info(request,'fil uploaded')
        return redirect('employee_overview',pk)
def shareemail(request,pk):
    try:
            if request.method == 'POST':
                emails_string = request.POST['email']

    
                emails_list = [email.strip() for email in emails_string.split(',')]
                print(emails_list)
                p=payroll_employee.objects.get(id=pk)
                        
                context = {'p':p}
                template_path = 'zohomodules/payroll-employee/mailoverview.html'
                template = get_template(template_path)
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'overview page - {p.id}.pdf'
                subject = f"overview page  - {p.first_name}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached employee details - File-{p.id}.\n--\nRegards,\n", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)
                messages.success(request, 'over view page has been shared via email successfully..!')
                return redirect('employee_overview',pk)
    except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect('employee_overview',pk)
    
